"""Excel writer module for YMOS — professional-grade Excel generation."""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# --- Style Constants ---

FONT_DATA = Font(name="Times New Roman", size=11)
FONT_TITLE = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Times New Roman", size=11, bold=True)
FONT_STAT = Font(name="Times New Roman", size=11, bold=True)

FILL_TITLE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_STAT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

BORDER_NONE = Border()
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# --- Data Validation ---

class ValidationError(Exception):
    """Raised when data validation fails."""
    pass


def validate_data(data: dict[str, Any], schema: dict[str, Any] | None = None) -> list[str]:
    """Validate data against an optional schema. Returns list of error messages."""
    errors: list[str] = []

    if schema is None:
        return errors

    required = schema.get("required", [])
    for field in required:
        if field not in data or data[field] is None:
            errors.append(f"缺少必填字段: {field}")

    types = schema.get("types", {})
    for field, expected_type in types.items():
        if field in data and data[field] is not None:
            if expected_type == "number" and not isinstance(data[field], (int, float)):
                errors.append(f"字段 {field} 应为数字类型")
            elif expected_type == "date" and not isinstance(data[field], (dt.date, dt.datetime)):
                errors.append(f"字段 {field} 应为日期类型")

    ranges = schema.get("ranges", {})
    for field, (min_val, max_val) in ranges.items():
        if field in data and data[field] is not None:
            val = data[field]
            if isinstance(val, (int, float)):
                if val < min_val or val > max_val:
                    errors.append(f"字段 {field} 值 {val} 超出范围 [{min_val}, {max_val}]")

    return errors


# --- Template Loading ---

import json


def load_template(template_path: Path) -> dict:
    """Load a JSON template file."""
    if not template_path.exists():
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    with open(template_path, "r", encoding="utf-8") as f:
        return json.load(f)


def resolve_template_path(template_name: str, template_dir: Path | None = None) -> Path:
    """Resolve template name to full path."""
    if template_dir is None:
        template_dir = Path(__file__).parent.parent / ".claude" / "skills" / "ymos-excel-output" / "templates"

    path = template_dir / f"{template_name}.json"
    if path.exists():
        return path

    raise FileNotFoundError(f"模板不存在: {template_name}")


# --- Core Writer ---

def write_excel(
    output_path: Path,
    sheets: list[dict[str, Any]],
    title: str | None = None,
    params: dict[str, str] | None = None,
) -> Path:
    """Main entry: generate Excel file from sheet definitions.

    Args:
        output_path: Output file path (.xlsx).
        sheets: List of sheet configs. Each config has:
            - name: str (sheet name)
            - title: str | None (sheet title row)
            - headers: list[str] (column headers)
            - rows: list[list[Any]] (data rows, supports formulas as strings starting with '=')
            - stat_rows: list[list[Any]] | None (summary rows)
            - column_widths: list[int] | None
            - comments: list[dict] | None — [{row, col, text, author}]
        title: Optional workbook title.
        params: Optional params for template substitution.

    Returns:
        Path to the generated file.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_cfg in sheets:
        ws = wb.create_sheet(title=sheet_cfg.get("name", "Sheet1"))
        row_idx = 1

        # Sheet title
        sheet_title = sheet_cfg.get("title")
        if sheet_title:
            headers = sheet_cfg.get("headers", [])
            col_count = len(headers) if headers else 1
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=col_count)
            cell = ws.cell(row=row_idx, column=1, value=sheet_title)
            cell.font = FONT_TITLE
            cell.fill = FILL_TITLE
            cell.alignment = ALIGN_CENTER
            row_idx += 1
            row_idx += 1  # blank row

        # Headers
        headers = sheet_cfg.get("headers", [])
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
        row_idx += 1

        data_start_row = row_idx

        # Data rows
        for row_data in sheet_cfg.get("rows", []):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = FONT_DATA
                cell.fill = FILL_WHITE
                cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER
                if isinstance(value, float):
                    cell.number_format = '#,##0.0'
            row_idx += 1

        data_end_row = row_idx - 1

        # Stat rows
        for stat_row in sheet_cfg.get("stat_rows", []):
            for col_idx, value in enumerate(stat_row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = FONT_STAT
                cell.fill = FILL_STAT
                cell.alignment = ALIGN_LEFT if col_idx == 1 else ALIGN_CENTER
                cell.border = THIN_BORDER
            row_idx += 1

        # Comments
        for comment_cfg in sheet_cfg.get("comments", []):
            r = comment_cfg["row"]
            c = comment_cfg["col"]
            text = comment_cfg["text"]
            author = comment_cfg.get("author", "YMOS")
            cell = ws.cell(row=r, column=c)
            cell.comment = openpyxl.comments.Comment(text, author)

        # Auto-add data source comments for non-formula cells
        source = sheet_cfg.get("data_source")
        if source:
            for r in range(data_start_row, data_end_row + 1):
                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith("=")):
                        if cell.comment is None:
                            cell.comment = openpyxl.comments.Comment(source, "YMOS")

        # Column widths
        widths = sheet_cfg.get("column_widths")
        if widths:
            for col_idx, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        else:
            for col_idx in range(1, len(headers) + 1):
                max_len = len(str(headers[col_idx - 1])) if headers else 10
                for r in range(data_start_row, data_end_row + 1):
                    val = ws.cell(row=r, column=col_idx).value
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    wb.save(str(output_path))
    return output_path


def write_excel_from_template(
    template_name: str,
    data: dict[str, Any],
    output_path: Path,
    template_dir: Path | None = None,
) -> Path:
    """Generate Excel from a named template with data binding."""
    template_path = resolve_template_path(template_name, template_dir)
    template = load_template(template_path)

    # Simple data binding: replace {key} in string values
    sheets = template.get("sheets", [])
    for sheet in sheets:
        for key, value in data.items():
            # Bind in title
            if "title" in sheet and isinstance(sheet["title"], str):
                sheet["title"] = sheet["title"].replace(f"{{{key}}}", str(value))
            # Bind in row values
            new_rows = []
            for row in sheet.get("rows", []):
                new_row = []
                for cell_val in row:
                    if isinstance(cell_val, str):
                        new_row.append(cell_val.replace(f"{{{key}}}", str(value)))
                    else:
                        new_row.append(cell_val)
                new_rows.append(new_row)
            sheet["rows"] = new_rows

    return write_excel(output_path, sheets)


def detect_circular_ref(formulas: dict[str, str]) -> list[str]:
    """Detect circular references in cell formulas. Returns list of circular cell addresses."""
    circular = []
    visited = set()

    def dfs(cell: str, path: set) -> None:
        if cell in path:
            circular.append(cell)
            return
        if cell in visited:
            return
        visited.add(cell)
        path.add(cell)
        formula = formulas.get(cell, "")
        if isinstance(formula, str) and formula.startswith("="):
            # Simple ref extraction
            import re
            refs = re.findall(r'([A-Z]+[0-9]+)', formula)
            for ref in refs:
                dfs(ref, path)
        path.discard(cell)

    for cell in formulas:
        dfs(cell, set())

    return circular

    return circular
